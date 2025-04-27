import xlrd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def read_excel_file(filename: str, sheet_names: tuple = None) -> {tuple}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    if not sheet_names:
        sheet_names = workbook.sheetnames
    all_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        all_data[sheet_name] = data
    workbook.close()
    return all_data


def save_excel_file(filename: str, data):
    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(filename)


def read_xls(input_file) -> dict:
    wb = xlrd.open_workbook(input_file)
    out_dict = {}
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for row in range(ws.nrows):
            row_value = []
            for col in range(ws.ncols):
                cell_value = ws.cell_value(row, col)
                row_value.append(cell_value)
            rows.append(row_value)
        out_dict[sheet_name] = rows
    return out_dict


